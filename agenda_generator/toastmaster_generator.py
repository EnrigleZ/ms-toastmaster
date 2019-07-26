import json
import openpyxl
import re
import datetime
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import sys
import os
import webbrowser


def min_distance(word1, word2):
    """
    :type word1: str
    :type word2: str
    :rtype: int
    """
    dp_matrix = list(range(0, len(word2) + 1))
    for end_index1 in range(1, len(word1) + 1):
        dp_row = [end_index1]
        dp_matrix.append(dp_row)
        for end_index2 in range(1, len(word2) + 1):
            candidate_list = [
                dp_matrix[end_index1][end_index2 - 1] + 1,
                dp_matrix[end_index1 - 1][end_index2] + 1
            ]
            # replace the last char in word1
            if word1[end_index1 - 1] == word2[end_index2 - 1]:
                candidate_list.append(dp_matrix[end_index1 - 1][end_index2 - 1])
            else:
                candidate_list.append(dp_matrix[end_index1 - 1][end_index2 - 1] + 1)
            dp_row.append(min(candidate_list))
    return dp_matrix[len(word1)][len(word2)]


def try_get_str(s):
    return str(s) if s is not None else ""


def cover_border(base_border, top=None, bottom=None, left=None, right=None):
    return Border(
        top=base_border.top if top is None else top,
        bottom=base_border.bottom if bottom is None else bottom,
        left=base_border.left if left is None else left,
        right=base_border.right if right is None else right,
    )


def style_range(ws, cell_range, border=Border()):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param cell_range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    """

    rows = ws[cell_range]
    for cell in rows[0]:
        cell.border = cover_border(cell.border, top=border.top)
    for cell in rows[-1]:
        cell.border = cover_border(cell.border, bottom=border.bottom)

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = cover_border(l.border, left=border.left)
        r.border = cover_border(r.border, right=border.right)


def set_active_sheet_by_name(wb, sheet_name):
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == sheet_name:
            wb.active = s
            return True
    return False


def get_meeting_date_str(next_meeting):
    return "{2}{0:02}{1:02}".format(
        next_meeting["month"],
        next_meeting["day"],
        datetime.datetime.now().year
    )


class ToastmasterAgendaGenerator:
    def __init__(self):
        self.roles = [
        ]

        self.pathways = [
        ]

        self.CCs = [
        ]

        self.time_dict = {
        }

    def read_info_from_call_role(self, call_role_text):
        meetings = []
        meeting_info = {}
        for line in call_role_text.split('\n'):
            if line.find("Call Role") is not -1:
                meetings.append(meeting_info)
                meeting_info = {}

            if line.find(":") is not -1:
                ti = line.find(":")
                role_name, member_name = line[:ti], line[ti+1:]
                meeting_info[role_name.strip()] = member_name.strip().strip("\ufe0f")
            else:
                line = line.replace(u"中文", "Chinese")
                m = re.match(r"([0-9]+)/([0-9]+) *\((Chinese|English)\)", line)
                if m is not None:
                    meeting_info["month"] = int(m.group(1))
                    meeting_info["day"] = int(m.group(2))
                    meeting_info["language"] = m.group(3)

        meeting_info["speech_count"] = 0
        for i in range(1, 4):
            if "SP{0}".format(i) in meeting_info and len(meeting_info["SP{0}".format(i)].strip()) is not 0:
                meeting_info["speech_count"] += 1

        meetings.append(meeting_info)
        meetings = list(filter(
            lambda x: len(x) is not 0,
            meetings
        ))
        return meetings

    def set_role(self, next_meeting, role_sheet, member_info_path):
        member_info_list = json.load(
            open(member_info_path, "r")
        )
        speech_levels = []
        for i in range(0, len(self.roles)):
            row_index = i + 2
            current_role = self.roles[i]
            if "nick" not in current_role:
                current_role["nick"] = current_role["name"]

            role_taker_name = ""
            if current_role['nick'] in next_meeting:
                role_taker_name = next_meeting[current_role['nick']]
            if len(role_taker_name) is 0 and "default_taker" in current_role:
                role_taker_name = current_role["default_taker"]

            role_taker = None
            if len(role_taker_name) is not 0:
                for member_info in member_info_list:
                    if member_info["English Name"].find(role_taker_name) is not -1 or \
                            member_info["Chinese Name"].find(role_taker_name) is not -1:
                        role_taker = member_info
                        break
            else:
                role_taker = {
                    "English Name": "TBD",
                    "Chinese Name": "TBD",
                    "CC Level": None,
                    "CC Level 2": None,
                    "Pathway Level": None,
                    "Speech Records": [],
                    "Role Records": [],
                }
            if role_taker is None:
                role_taker = {
                    "English Name": role_taker_name,
                    "Chinese Name": role_taker_name,
                    "CC Level": None,
                    "CC Level 2": None,
                    "Pathway Level": None,
                    "Speech Records": [],
                    "Role Records": [],
                }
            if role_taker is not None:
                is_on_pathway = role_taker["Pathway Level"] is not None or role_taker["CC Level"] is None

                name = role_taker["{0} Name".format(next_meeting["language"])]
                role_sheet['B{0}'.format(row_index)] = name
                role_sheet['C{0}'.format(row_index)] = try_get_str(role_taker["Pathway Level"]) \
                    if is_on_pathway else try_get_str(role_taker["CC Level"])
                role_sheet['D{0}'.format(row_index)] = try_get_str(role_taker["CC Level 2"])

                if current_role["name"].find("Speaker") is 0:
                    if is_on_pathway is True:
                        current_level = self.pathways[self.pathways.index(role_taker["Pathway Level"])+1]
                        role_taker["Pathway Level"] = current_level
                    else:
                        current_level = self.CCs[self.CCs.index(role_taker["CC Level"]) + 1]
                        role_taker["CC Level"] = current_level
                    role_taker["Speech Records"].append({
                        "Level": current_level,
                        "Date": "{2}{0:02}{1:02}".format(
                            next_meeting["month"],
                            next_meeting["day"],
                            datetime.datetime.now().year
                        )
                    })
                    speech_levels.append(current_level)
                elif current_role["name"] not in ["SAA", "President"]:
                    role_taker["Role Records"].append({
                        "Role": current_role["name"],
                        "Date": "{2}{0:02}{1:02}".format(
                            next_meeting["month"],
                            next_meeting["day"],
                            datetime.datetime.now().year
                        )
                    })
        json.dump(
            member_info_list,
            open("{0}.member_info.json".format(
                get_meeting_date_str(next_meeting)
            ), "w", encoding="utf-8"),
            indent=2
        )
        return speech_levels

    def load_settings(self, settings_path):
        settings = json.load(open(settings_path, "r", encoding="utf-8"))
        for var in settings:
            setattr(self, var, settings[var])

    def generate_agenda(self, origin_text, current_log_path):
        for next_meeting in self.read_info_from_call_role(origin_text):
            open(
                "{0}.call_role.txt".format(get_meeting_date_str(next_meeting)),
                "w",
                encoding="utf-8"
            ).write(origin_text)
            print(json.dumps(next_meeting, indent=2))
            xlsx_template = openpyxl.load_workbook("ToastMaster_Template.xlsx")
            role_sheet = xlsx_template["Roles"]

            speech_levels = self.set_role(next_meeting, role_sheet, current_log_path)

            is_english = next_meeting["language"] == "English"
            agenda_sheet_prefix = "Agenda" if is_english else "Chinese Agenda"
            speech_count = next_meeting["speech_count"]
            agenda_sheet_name = "{0}-{1}".format(agenda_sheet_prefix, speech_count)
            agenda_sheet = xlsx_template[agenda_sheet_name]
            set_active_sheet_by_name(xlsx_template, agenda_sheet_name)

            speech_rows = [27, 29, 31]
            if is_english is True:
                agenda_sheet["A8"] = "Theme Today: {0}".format(next_meeting["Theme"])
            else:
                agenda_sheet["A7"] = "本期主题:  “{0}”".format(next_meeting["Theme"])

            for i in range(0, speech_count):
                agenda_sheet["C{0}".format(speech_rows[i])] = next_meeting["SP{0} Topic".format(i + 1)]
                current_level = speech_levels[i]
                if current_level not in self.time_dict:
                    self.time_dict[current_level] = ["0:06", "4-6", 4, 5, 6]
                duration, time_range, green_time, yellow_time, red_time = self.time_dict[current_level]
                reorg = [time_range, green_time, yellow_time, red_time, duration]
                for j in range(0, len(reorg)):
                    agenda_sheet["{0}{1}".format(chr(ord('G') + j), speech_rows[i])] = reorg[j]

            icon_img = Image('Icon.png')
            icon_img.anchor = 'A1'
            agenda_sheet.add_image(icon_img)

            qr_img = Image('QR Code.png')
            qr_img.anchor = 'G1'
            agenda_sheet.add_image(qr_img)

            qr_img = Image('qrcode-vote-{0}.png'.format(speech_count))
            qr_img.anchor = 'E{0}'.format(4 * speech_count + 34)
            agenda_sheet.add_image(qr_img)

            side = Side(border_style="medium", color='000000')
            border = Border(
                left=side,
                right=side,
                top=side,
                bottom=side,
            )
            style_range(agenda_sheet, 'A1:J3', border)
            style_range(agenda_sheet, 'A4:J5', border)
            style_range(agenda_sheet, 'A8:J8', border)
            style_range(agenda_sheet, 'A20:J20', border)
            style_range(agenda_sheet, 'A25:J25', border)
            style_range(agenda_sheet, 'A{0}:J{0}'.format(26 + 3 * speech_count), border)

            xlsx_template.save("{0}.agenda.xlsx".format(
                get_meeting_date_str(next_meeting)
            ))


def __main__():
    working_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(working_dir)
    if len(sys.argv) != 3:
        last_date = input("Please enter last meeting date (date for latest member info):\n")
        current_log_path = "{0}.member_info.json".format(last_date)
        call_role_path = input("Please enter path for role calling:\n")
    else:
        _, current_log_path, call_role_path = sys.argv

    # read call role file
    origin_text = open(call_role_path, "r", encoding="utf-8").read()

    generator = ToastmasterAgendaGenerator()

    # read setting
    generator.load_settings("settings.json")

    generator.generate_agenda(origin_text, current_log_path)

    webbrowser.open("https://wj.qq.com/mine.html")

    input("successfully generated.")


if __name__ == "__main__":
    __main__()